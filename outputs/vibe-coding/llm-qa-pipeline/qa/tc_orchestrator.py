"""자동 QA 사이클 관리 및 오케스트레이션 모듈"""

import os
from datetime import datetime
from typing import List, Dict, Tuple, Any
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from .tc_generator import generate_tcs, TC_TYPE_DESCRIPTIONS
from .tc_validator import batch_validate, filter_valid_tcs
from .tc_regenerator import regenerate_from_failures, print_regeneration_summary
from .tc_heuristic import (
    generate_targeted_tcs,
    analyze_patterns,
    record_execution_result,
    print_pattern_analysis
)
from .qa_pipeline import run_tc
from .analytics import collect_all_results, analyze_failures


def generate_all_types(
    n_per_type: int = 2,
    validate: bool = True
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """
    모든 10개 유형에서 각각 n개씩 TC 생성

    Args:
        n_per_type: 각 유형당 생성 개수
        validate: 자동 검증 포함 여부

    Returns:
        (tc_list, summary)
        summary: {
            "total": int,
            "by_type": {type: count},
            "validation": {valid: int, invalid: int, duplicate: int}
        }
    """
    print(f"\n{'='*60}")
    print(f"🚀 모든 유형 TC 생성 시작")
    print(f"{'='*60}")
    print(f"   유형 수: {len(TC_TYPE_DESCRIPTIONS)}")
    print(f"   각 유형당: {n_per_type}개")

    all_tcs = []
    by_type_count = {}

    for idx, tc_type in enumerate(TC_TYPE_DESCRIPTIONS.keys(), 1):
        try:
            tcs = generate_tcs(tc_type, n_per_type)
            all_tcs.extend(tcs)
            by_type_count[tc_type] = len(tcs)
            print(f"   [{idx:2d}] {tc_type:10s}: {len(tcs)}개 생성")
        except Exception as e:
            by_type_count[tc_type] = 0
            print(f"   [{idx:2d}] {tc_type:10s}: ⚠️  생성 실패 - {str(e)[:40]}")

    # 검증
    validation_summary = {"valid": 0, "invalid": 0, "duplicate": 0}

    if validate and all_tcs:
        print(f"\n   📊 생성된 TC 검증 중...")
        validation_result = batch_validate(all_tcs)
        validation_summary = {
            "valid": validation_result["valid"],
            "invalid": validation_result["invalid"],
            "duplicate": validation_result["duplicates"]
        }
        print(f"      유효: {validation_summary['valid']}/{validation_result['total']}")

        # 유효한 TC만 필터링
        all_tcs, _ = filter_valid_tcs(all_tcs)

    summary = {
        "total": len(all_tcs),
        "by_type": by_type_count,
        "validation": validation_summary
    }

    return all_tcs, summary


def auto_qa_cycle(
    num_iterations: int = 2,
    n_per_type: int = 2,
    use_heuristics: bool = False,
    verbose: bool = True
) -> Dict[str, Any]:
    """
    자동 QA 사이클 실행 (여러 반복)

    Args:
        num_iterations: 반복 횟수 (1~3)
        n_per_type: 각 유형당 TC 개수
        use_heuristics: 휴리스틱 기반 생성 사용 여부
        verbose: 상세 출력

    Returns:
        {
            "summary": {...},
            "iterations": [...],
            "patterns": {...}
        }
    """
    if num_iterations < 1 or num_iterations > 3:
        raise ValueError("num_iterations는 1~3 범위여야 합니다")

    start_time = datetime.now()
    cycle_result = {
        "summary": {
            "start_time": start_time.isoformat(),
            "end_time": None,
            "duration_minutes": 0,
            "total_iterations": num_iterations,
            "total_tcs_generated": 0,
            "total_tcs_executed": 0,
            "overall_pass_rate": 0.0,
            "overall_fail_count": 0
        },
        "iterations": [],
        "patterns": {}
    }

    total_generated = 0
    total_executed = 0
    total_passes = 0
    total_checks = 0

    # ===== Iteration 1: all_types 전략 =====
    iteration_result = {
        "iteration": 1,
        "strategy": "all_types",
        "timestamp": datetime.now().isoformat(),
        "generation": {},
        "execution": {},
        "top_failures": []
    }

    print(f"\n{'='*60}")
    print(f"📍 Iteration 1: all_types 전략")
    print(f"{'='*60}")

    tcs, gen_summary = generate_all_types(n_per_type=n_per_type, validate=True)
    iteration_result["generation"] = {
        "strategy": "all_types",
        "n_per_type": n_per_type,
        "tcs_generated": gen_summary["total"],
        "valid_tcs": gen_summary["validation"]["valid"],
        "invalid_tcs": gen_summary["validation"]["invalid"],
        "validation_pass_rate": gen_summary["validation"]["valid"] / gen_summary["total"] if gen_summary["total"] > 0 else 0.0
    }

    total_generated += gen_summary["total"]

    # TC 실행
    if tcs:
        print(f"\n   🔄 생성된 {len(tcs)}개 TC 실행 중...")
        exec_results = []
        for idx, tc in enumerate(tcs, 1):
            if verbose and idx % 5 == 0:
                print(f"      진행: {idx}/{len(tcs)}")
            try:
                result = run_tc(
                    tc["tc_id"],
                    tc["user_input"],
                    tc.get("slide_count", "auto"),
                    tc.get("expected_pages"),
                    tc.get("expected_language", "ko")
                )
                exec_results.append(result["result"])

                # 패턴 기록
                tc_type = tc.get("tc_type", "unknown")
                for section in ["rule", "llm_story", "llm_text"]:
                    if section in result["result"] and result["result"][section]:
                        record_execution_result(tc_type, result["result"][section])

            except Exception as e:
                if verbose:
                    print(f"      ⚠️  {tc['tc_id']}: 실행 실패 - {str(e)[:30]}")

        total_executed += len(exec_results)

        # 분석
        if exec_results:
            analysis = analyze_failures(exec_results)
            pass_rate = analysis.get("overall_pass_rate", 0.0)
            fail_count = analysis.get("fail_count", 0)
            total_checks += analysis.get("total_checks", 0)
            total_passes += analysis.get("pass_count", 0)

            iteration_result["execution"] = {
                "tcs_executed": len(exec_results),
                "pass_rate": pass_rate,
                "pass_count": analysis.get("pass_count", 0),
                "fail_count": fail_count,
                "skip_count": analysis.get("skip_count", 0)
            }

            # 상위 FAIL
            top_fails = sorted(
                analysis.get("failures_by_code", {}).items(),
                key=lambda x: x[1],
                reverse=True
            )[:5]
            iteration_result["top_failures"] = [
                {"code": code, "count": count, "percentage": round(count / total_checks * 100, 1)}
                for code, count in top_fails
            ]

            print(f"   ✅ 실행 완료: {pass_rate*100:.1f}% PASS")

    cycle_result["iterations"].append(iteration_result)

    # ===== Iteration 2: failure_based 전략 =====
    if num_iterations >= 2:
        iteration_result = {
            "iteration": 2,
            "strategy": "failure_based",
            "timestamp": datetime.now().isoformat(),
            "generation": {},
            "execution": {},
            "improvement": {},
            "top_failures": []
        }

        print(f"\n{'='*60}")
        print(f"📍 Iteration 2: failure_based 전략")
        print(f"{'='*60}")

        # FAIL 기반 재생성
        try:
            if exec_results:
                analysis = analyze_failures(exec_results)
                new_tcs = regenerate_from_failures(
                    analysis,
                    n_per_code=3,
                    max_codes=5
                )

                if new_tcs:
                    iteration_result["generation"] = {
                        "strategy": "failure_based",
                        "tcs_generated": len(new_tcs),
                        "target_codes": len(analysis.get("failures_by_code", {}))
                    }

                    total_generated += len(new_tcs)

                    # 실행
                    print(f"\n   🔄 생성된 {len(new_tcs)}개 TC 실행 중...")
                    new_exec_results = []
                    for idx, tc in enumerate(new_tcs, 1):
                        if verbose and idx % 5 == 0:
                            print(f"      진행: {idx}/{len(new_tcs)}")
                        try:
                            result = run_tc(
                                tc["tc_id"],
                                tc["user_input"],
                                tc.get("slide_count", "auto"),
                                tc.get("expected_pages"),
                                tc.get("expected_language", "ko")
                            )
                            new_exec_results.append(result["result"])

                            # 패턴 기록
                            tc_type = tc.get("tc_type", "unknown")
                            for section in ["rule", "llm_story", "llm_text"]:
                                if section in result["result"] and result["result"][section]:
                                    record_execution_result(tc_type, result["result"][section])

                        except Exception as e:
                            if verbose:
                                print(f"      ⚠️  {tc['tc_id']}: 실행 실패")

                    total_executed += len(new_exec_results)

                    # 분석
                    if new_exec_results:
                        new_analysis = analyze_failures(new_exec_results)
                        new_pass_rate = new_analysis.get("overall_pass_rate", 0.0)
                        new_fail_count = new_analysis.get("fail_count", 0)
                        total_checks += new_analysis.get("total_checks", 0)
                        total_passes += new_analysis.get("pass_count", 0)

                        iteration_result["execution"] = {
                            "tcs_executed": len(new_exec_results),
                            "pass_rate": new_pass_rate,
                            "pass_count": new_analysis.get("pass_count", 0),
                            "fail_count": new_fail_count,
                            "skip_count": new_analysis.get("skip_count", 0)
                        }

                        # 개선도 계산
                        if "execution" in cycle_result["iterations"][0]:
                            prev_pass_rate = cycle_result["iterations"][0]["execution"].get("pass_rate", 0.0)
                            improvement = new_pass_rate - prev_pass_rate
                            iteration_result["improvement"] = {
                                "pass_rate_delta": round(improvement, 3),
                                "direction": "↑" if improvement > 0 else "↓" if improvement < 0 else "→"
                            }

                        # 상위 FAIL
                        top_fails = sorted(
                            new_analysis.get("failures_by_code", {}).items(),
                            key=lambda x: x[1],
                            reverse=True
                        )[:5]
                        iteration_result["top_failures"] = [
                            {"code": code, "count": count}
                            for code, count in top_fails
                        ]

                        print(f"   ✅ 실행 완료: {new_pass_rate*100:.1f}% PASS")

        except Exception as e:
            print(f"   ⚠️  Iteration 2 실패: {str(e)[:50]}")

        cycle_result["iterations"].append(iteration_result)

    # ===== Iteration 3: heuristic 전략 (선택) =====
    if num_iterations >= 3 and use_heuristics:
        iteration_result = {
            "iteration": 3,
            "strategy": "heuristic",
            "timestamp": datetime.now().isoformat(),
            "generation": {},
            "execution": {},
            "top_failures": []
        }

        print(f"\n{'='*60}")
        print(f"📍 Iteration 3: heuristic 전략")
        print(f"{'='*60}")

        try:
            # 휴리스틱 기반 생성
            heur_tcs = generate_targeted_tcs(
                n_per_pattern=2,
                top_patterns=5
            )

            if heur_tcs:
                iteration_result["generation"] = {
                    "strategy": "heuristic",
                    "tcs_generated": len(heur_tcs)
                }

                total_generated += len(heur_tcs)

                # 실행
                print(f"\n   🔄 생성된 {len(heur_tcs)}개 TC 실행 중...")
                heur_exec_results = []
                for idx, tc in enumerate(heur_tcs, 1):
                    if verbose and idx % 5 == 0:
                        print(f"      진행: {idx}/{len(heur_tcs)}")
                    try:
                        result = run_tc(
                            tc["tc_id"],
                            tc["user_input"],
                            tc.get("slide_count", "auto"),
                            tc.get("expected_pages"),
                            tc.get("expected_language", "ko")
                        )
                        heur_exec_results.append(result["result"])

                        # 패턴 기록
                        tc_type = tc.get("tc_type", "unknown")
                        for section in ["rule", "llm_story", "llm_text"]:
                            if section in result["result"] and result["result"][section]:
                                record_execution_result(tc_type, result["result"][section])

                    except Exception as e:
                        if verbose:
                            print(f"      ⚠️  {tc['tc_id']}: 실행 실패")

                total_executed += len(heur_exec_results)

                # 분석
                if heur_exec_results:
                    heur_analysis = analyze_failures(heur_exec_results)
                    heur_pass_rate = heur_analysis.get("overall_pass_rate", 0.0)
                    heur_fail_count = heur_analysis.get("fail_count", 0)
                    total_checks += heur_analysis.get("total_checks", 0)
                    total_passes += heur_analysis.get("pass_count", 0)

                    iteration_result["execution"] = {
                        "tcs_executed": len(heur_exec_results),
                        "pass_rate": heur_pass_rate,
                        "pass_count": heur_analysis.get("pass_count", 0),
                        "fail_count": heur_fail_count,
                        "skip_count": heur_analysis.get("skip_count", 0)
                    }

                    # 상위 FAIL
                    top_fails = sorted(
                        heur_analysis.get("failures_by_code", {}).items(),
                        key=lambda x: x[1],
                        reverse=True
                    )[:5]
                    iteration_result["top_failures"] = [
                        {"code": code, "count": count}
                        for code, count in top_fails
                    ]

                    print(f"   ✅ 실행 완료: {heur_pass_rate*100:.1f}% PASS")

        except Exception as e:
            print(f"   ⚠️  Iteration 3 실패: {str(e)[:50]}")

        cycle_result["iterations"].append(iteration_result)

    # ===== 최종 요약 =====
    end_time = datetime.now()
    duration = (end_time - start_time).total_seconds() / 60

    cycle_result["summary"]["end_time"] = end_time.isoformat()
    cycle_result["summary"]["duration_minutes"] = round(duration, 1)
    cycle_result["summary"]["total_tcs_generated"] = total_generated
    cycle_result["summary"]["total_tcs_executed"] = total_executed
    cycle_result["summary"]["overall_pass_rate"] = (
        total_passes / total_checks if total_checks > 0 else 0.0
    )
    cycle_result["summary"]["overall_fail_count"] = total_checks - total_passes

    # 패턴 분석
    try:
        patterns = analyze_patterns(min_samples=1)
        cycle_result["patterns"] = {
            "top_patterns": [
                {"pattern": f"{t}|{c}", "fail_ratio": f"{r*100:.1f}%", "samples": s}
                for t, c, r, s in patterns[:10]
            ],
            "pattern_count": len(patterns)
        }
    except:
        cycle_result["patterns"] = {}

    print(f"\n{'='*60}")
    print(f"✅ QA 사이클 완료")
    print(f"{'='*60}")
    print(f"   총 생성: {total_generated}개")
    print(f"   총 실행: {total_executed}개")
    print(f"   전체 PASS율: {cycle_result['summary']['overall_pass_rate']*100:.1f}%")
    print(f"   소요 시간: {duration:.1f}분")
    print(f"{'='*60}\n")

    return cycle_result


def save_cycle_report(
    cycle_result: Dict[str, Any],
    filename: str = "qa_cycle_report.xlsx",
    output_dir: str = "./reports"
) -> str:
    """
    최종 QA 사이클 결과를 Excel로 저장

    Args:
        cycle_result: auto_qa_cycle() 반환값
        filename: 파일명
        output_dir: 출력 디렉토리

    Returns:
        저장된 파일 경로
    """
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    # ===== 시트 1: 요약 =====
    ws_summary = wb.create_sheet("요약", 0)
    summary = cycle_result.get("summary", {})

    ws_summary["A1"] = "QA 사이클 요약"
    ws_summary["A1"].font = Font(bold=True, size=14)

    row = 3
    ws_summary[f"A{row}"] = "항목"
    ws_summary[f"B{row}"] = "값"
    row += 1

    summary_items = [
        ("시작 시간", summary.get("start_time", "").split("T")[0]),
        ("반복 횟수", summary.get("total_iterations", 0)),
        ("생성 TC", summary.get("total_tcs_generated", 0)),
        ("실행 TC", summary.get("total_tcs_executed", 0)),
        ("전체 PASS율", f"{summary.get('overall_pass_rate', 0)*100:.1f}%"),
        ("소요 시간", f"{summary.get('duration_minutes', 0):.1f}분"),
    ]

    for label, value in summary_items:
        ws_summary[f"A{row}"] = label
        ws_summary[f"B{row}"] = value
        row += 1

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 30

    # ===== 시트 2: 반복별 결과 =====
    ws_iterations = wb.create_sheet("반복별 결과", 1)

    ws_iterations["A1"] = "반복별 QA 결과"
    ws_iterations["A1"].font = Font(bold=True, size=14)

    row = 3
    iterations = cycle_result.get("iterations", [])

    for iteration in iterations:
        iter_num = iteration.get("iteration", 0)
        strategy = iteration.get("strategy", "")

        ws_iterations[f"A{row}"] = f"Iteration {iter_num} ({strategy})"
        ws_iterations[f"A{row}"].font = Font(bold=True)
        row += 1

        gen = iteration.get("generation", {})
        exe = iteration.get("execution", {})

        ws_iterations[f"A{row}"] = "생성 TC"
        ws_iterations[f"B{row}"] = gen.get("tcs_generated", 0)
        row += 1

        ws_iterations[f"A{row}"] = "실행 TC"
        ws_iterations[f"B{row}"] = exe.get("tcs_executed", 0)
        row += 1

        ws_iterations[f"A{row}"] = "PASS율"
        ws_iterations[f"B{row}"] = f"{exe.get('pass_rate', 0)*100:.1f}%"
        row += 1

        row += 1  # 빈 줄

    ws_iterations.column_dimensions["A"].width = 25
    ws_iterations.column_dimensions["B"].width = 20

    # ===== 시트 3: 패턴 분석 =====
    ws_patterns = wb.create_sheet("패턴", 2)

    ws_patterns["A1"] = "발견된 FAIL 패턴"
    ws_patterns["A1"].font = Font(bold=True, size=14)

    patterns = cycle_result.get("patterns", {})
    top_patterns = patterns.get("top_patterns", [])

    if top_patterns:
        row = 3
        ws_patterns[f"A{row}"] = "패턴"
        ws_patterns[f"B{row}"] = "FAIL율"
        ws_patterns[f"C{row}"] = "샘플수"
        row += 1

        for pattern in top_patterns:
            ws_patterns[f"A{row}"] = pattern.get("pattern", "")
            ws_patterns[f"B{row}"] = pattern.get("fail_ratio", "")
            ws_patterns[f"C{row}"] = pattern.get("samples", 0)
            row += 1

    ws_patterns.column_dimensions["A"].width = 20
    ws_patterns.column_dimensions["B"].width = 15
    ws_patterns.column_dimensions["C"].width = 15

    # 저장
    wb.save(filepath)

    print(f"✅ 리포트 저장: {filepath}\n")
    return filepath
