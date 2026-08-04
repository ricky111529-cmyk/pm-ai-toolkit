import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RULE_CODES = ["R-01", "R-02", "R-03/R-10", "R-04", "R-06", "R-07", "R-08", "R-09"]
LLM_STORY_CODES = [f"L-{i:02d}" for i in range(1, 13)]
LLM_TEXT_CODES = [f"T-{i:02d}" for i in range(1, 6)] 
ALL_CODES = RULE_CODES + LLM_STORY_CODES + LLM_TEXT_CODES

FILL_PASS = PatternFill("solid", fgColor="C6EFCE")
FILL_FAIL = PatternFill("solid", fgColor="FFC7CE")
FILL_SKIP = PatternFill("solid", fgColor="F2F2F2")
FILL_HEADER = PatternFill("solid", fgColor="4472C4")
FILL_SECTION_RULE = PatternFill("solid", fgColor="D9E1F2")
FILL_SECTION_LLM = PatternFill("solid", fgColor="E2EFDA")
FILL_SECTION_TEXT = PatternFill("solid", fgColor="FFF2CC")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _cell_value(results: dict, code: str) -> str:
    """결과 dict에서 코드에 해당하는 PASS/FAIL/SKIP 반환"""
    if not results or code not in results:
        return "SKIP"
    val = results[code]
    if isinstance(val, dict):
        return "PASS" if val.get("pass") else "FAIL"
    return "SKIP"


def _cell_fill(value: str) -> PatternFill:
    if value == "PASS":
        return FILL_PASS
    if value == "FAIL":
        return FILL_FAIL
    return FILL_SKIP


def _overall(row_values: list) -> str:
    """SKIP 제외, FAIL이 하나라도 있으면 FAIL, 전부 PASS면 PASS"""
    active = [v for v in row_values if v != "SKIP"]
    if not active:
        return "SKIP"
    return "FAIL" if "FAIL" in active else "PASS"


def save_excel(all_results: list, output_dir: str = ".") -> str:
    """
    all_results: run_tc / run_tc_multiturn 에서 반환한 dict 리스트
      각 dict 구조:
        {
          "tc_id": str,
          "turn": int,           # 단일턴=1, 멀티턴=1,2,3...
          "rule": dict,          # run_all_checks() 반환값
          "llm_story": dict,     # run_llm_checks_storyline() 반환값
          "llm_text": dict,      # run_llm_checks_text() 반환값 (없으면 None)
        }
    """
    wb = Workbook()

    # ── 시트 1: 상세 결과 ──────────────────────────────────────────────
    ws_detail = wb.active
    ws_detail.title = "상세 결과"

    # 헤더 행 1: 섹션 레이블
    section_headers = (
        [("", 1), ("", 1), ("Rule-based", len(RULE_CODES)),
         ("LLM — 스토리라인", len(LLM_STORY_CODES)),
         ("LLM — 텍스트 응답", len(LLM_TEXT_CODES)),
         ("", 1)]
    )
    col = 1
    for label, span in section_headers:
        cell = ws_detail.cell(row=1, column=col, value=label)
        if label:
            cell.font = Font(bold=True, color="FFFFFF")
            if label.startswith("Rule"):
                cell.fill = FILL_SECTION_RULE
            elif label.startswith("LLM — 스"):
                cell.fill = FILL_SECTION_LLM
            else:
                cell.fill = FILL_SECTION_TEXT
            cell.alignment = Alignment(horizontal="center")
            if span > 1:
                ws_detail.merge_cells(
                    start_row=1, start_column=col,
                    end_row=1, end_column=col + span - 1
                )
        col += span

    # 헤더 행 2: 컬럼명
    columns = ["TC ID", "Turn"] + ALL_CODES + ["Overall"]
    for c, name in enumerate(columns, 1):
        cell = ws_detail.cell(row=2, column=c, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # 데이터 행
    for r_idx, res in enumerate(all_results, 3):
        row_values = []
        for code in ALL_CODES:
            if code in RULE_CODES:
                row_values.append(_cell_value(res.get("rule"), code))
            elif code in LLM_STORY_CODES:
                row_values.append(_cell_value(res.get("llm_story"), code))
            else:
                row_values.append(_cell_value(res.get("llm_text"), code))

        overall = _overall(row_values)
        row_data = [res["tc_id"], res.get("turn", 1)] + row_values + [overall]

        for c_idx, val in enumerate(row_data, 1):
            cell = ws_detail.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            if isinstance(val, str) and val in ("PASS", "FAIL", "SKIP"):
                cell.fill = _cell_fill(val)

    # 열 너비 조정
    ws_detail.column_dimensions["A"].width = 12
    ws_detail.column_dimensions["B"].width = 6
    for c in range(3, len(columns) + 1):
        ws_detail.column_dimensions[get_column_letter(c)].width = 10
    ws_detail.freeze_panes = "C3"

    # ── 시트 2: 요약 ──────────────────────────────────────────────────
    ws_summary = wb.create_sheet("요약")

    # 전체 통계
    total = len(all_results)
    pass_count = sum(
        1 for res in all_results
        if _overall([
            _cell_value(res.get("rule"), c) for c in RULE_CODES
        ] + [
            _cell_value(res.get("llm_story"), c) for c in LLM_STORY_CODES
        ] + [
            _cell_value(res.get("llm_text"), c) for c in LLM_TEXT_CODES
        ]) == "PASS"
    )

    ws_summary.cell(row=1, column=1, value="전체 실행 수").font = Font(bold=True)
    ws_summary.cell(row=1, column=2, value=total)
    ws_summary.cell(row=2, column=1, value="전체 PASS").font = Font(bold=True)
    ws_summary.cell(row=2, column=2, value=pass_count)
    ws_summary.cell(row=3, column=1, value="전체 FAIL").font = Font(bold=True)
    ws_summary.cell(row=3, column=2, value=total - pass_count)
    rate = f"{pass_count / total * 100:.1f}%" if total else "-"
    ws_summary.cell(row=4, column=1, value="전체 PASS율").font = Font(bold=True)
    ws_summary.cell(row=4, column=2, value=rate)

    # 항목별 통계
    ws_summary.cell(row=6, column=1, value="항목").font = Font(bold=True)
    ws_summary.cell(row=6, column=2, value="총 실행").font = Font(bold=True)
    ws_summary.cell(row=6, column=3, value="PASS").font = Font(bold=True)
    ws_summary.cell(row=6, column=4, value="FAIL").font = Font(bold=True)
    ws_summary.cell(row=6, column=5, value="PASS율").font = Font(bold=True)
    for c in range(1, 6):
        ws_summary.cell(row=6, column=c).fill = FILL_HEADER
        ws_summary.cell(row=6, column=c).font = Font(bold=True, color="FFFFFF")
        ws_summary.cell(row=6, column=c).alignment = Alignment(horizontal="center")

    for r_idx, code in enumerate(ALL_CODES, 7):
        if code in RULE_CODES:
            vals = [_cell_value(res.get("rule"), code) for res in all_results]
        elif code in LLM_STORY_CODES:
            vals = [_cell_value(res.get("llm_story"), code) for res in all_results]
        else:
            vals = [_cell_value(res.get("llm_text"), code) for res in all_results]

        active = [v for v in vals if v != "SKIP"]
        p = active.count("PASS")
        f = active.count("FAIL")
        n = len(active)
        code_rate = f"{p / n * 100:.1f}%" if n else "-"

        ws_summary.cell(row=r_idx, column=1, value=code).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=r_idx, column=2, value=n).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=r_idx, column=3, value=p).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=r_idx, column=4, value=f).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=r_idx, column=5, value=code_rate).alignment = Alignment(horizontal="center")
        if f > 0:
            ws_summary.cell(row=r_idx, column=4).fill = FILL_FAIL

    for c in range(1, 6):
        ws_summary.column_dimensions[get_column_letter(c)].width = 14

    # 파일 저장
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"qa_report_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"\n리포트 저장 완료: {filepath}")
    return filepath


def print_summary(all_results: list):
    """터미널에 요약 출력"""
    total = len(all_results)
    if not total:
        print("결과 없음")
        return

    print("\n" + "=" * 60)
    print("QA 결과 요약")
    print("=" * 60)

    # 전체 PASS율
    overall_list = []
    for res in all_results:
        vals = (
            [_cell_value(res.get("rule"), c) for c in RULE_CODES] +
            [_cell_value(res.get("llm_story"), c) for c in LLM_STORY_CODES] +
            [_cell_value(res.get("llm_text"), c) for c in LLM_TEXT_CODES]
        )
        overall_list.append(_overall(vals))

    pass_count = overall_list.count("PASS")
    print(f"전체: {total}건  PASS: {pass_count}건  FAIL: {total - pass_count}건  "
          f"PASS율: {pass_count / total * 100:.1f}%")

    # TC별 요약
    print("\n[TC별 결과]")
    for res, ov in zip(all_results, overall_list):
        turn_str = f" (턴 {res['turn']})" if res.get("turn", 1) > 1 or "M" in res["tc_id"] else ""
        print(f"  {res['tc_id']}{turn_str}: {ov}")

    # 항목별 FAIL 빈도 (FAIL 있는 것만)
    print("\n[항목별 FAIL 빈도]")
    has_fail = False
    for code in ALL_CODES:
        if code in RULE_CODES:
            vals = [_cell_value(res.get("rule"), code) for res in all_results]
        elif code in LLM_STORY_CODES:
            vals = [_cell_value(res.get("llm_story"), code) for res in all_results]
        else:
            vals = [_cell_value(res.get("llm_text"), code) for res in all_results]
        active = [v for v in vals if v != "SKIP"]
        f = active.count("FAIL")
        if f > 0:
            n = len(active)
            print(f"  {code}: FAIL {f}/{n} ({f / n * 100:.1f}%)")
            has_fail = True
    if not has_fail:
        print("  없음")

    print("=" * 60)
