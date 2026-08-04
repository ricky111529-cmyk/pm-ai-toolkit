"""
Phase 6: QA 결과 분석 & 자동 리포팅

기능:
  - 전체 TC 자동 수집 및 실행
  - PASS/FAIL 빈도 분석
  - 항목별 FAIL 원인 분류
  - 카테고리별 성과 분석
  - Excel/HTML 리포트 생성
"""

from collections import Counter, defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from .reporter import RULE_CODES, LLM_STORY_CODES, LLM_TEXT_CODES, ALL_CODES


def collect_all_results(tc_list=None, verbose=True):
    """
    전체 TC 실행 및 결과 수집

    Args:
        tc_list: ["TC-001", "TC-002", ...] or ["TC-M01", ...] or None (모두)
        verbose: 진행 상황 출력

    Returns:
        list of {tc_id, turn, rule, llm_story, llm_text}
    """
    from .qa_pipeline import build_tc_map

    if tc_list is None:
        # 모든 TC 포함
        tc_list = [f"TC-{i:03d}" for i in range(1, 13)] + [f"TC-M{i:02d}" for i in range(1, 6)]

    # TC 맵 로드
    tc_map = build_tc_map()

    all_results = []
    for tc_id in tc_list:
        try:
            if verbose:
                print(f"[실행중] {tc_id}...", end=" ", flush=True)

            if tc_id in tc_map:
                # lambda 함수 실행
                result = tc_map[tc_id]()

                # 멀티턴은 리스트, 단일턴은 dict
                if isinstance(result, list):
                    all_results.extend(result)  # 리스트 평탄화
                else:
                    all_results.append(result)

                if verbose:
                    print("✓")
            else:
                if verbose:
                    print(f"✗ (TC 정의 없음)")
        except Exception as e:
            if verbose:
                print(f"✗ ({e})")

    return all_results


def analyze_failures(all_results):
    """
    FAIL 항목 분석

    Returns:
        {
          "overall_pass_rate": 0.85,
          "total_checks": 425,
          "pass_count": 361,
          "fail_count": 64,
          "skip_count": 0,
          "failures_by_code": {
            "R-01": 2,
            "L-05": 8,
            ...
          },
          "failures_by_category": {
            "Rule-based": {"total": 8, "codes": {...}},
            "LLM-Story": {"total": 32, "codes": {...}},
            "LLM-Text": {"total": 24, "codes": {...}}
          }
        }
    """
    failure_counter = Counter()
    category_failures = defaultdict(lambda: defaultdict(int))

    total_checks = 0
    pass_count = 0
    fail_count = 0
    skip_count = 0

    for result in all_results:
        # Rule-based
        for code in RULE_CODES:
            rule_res = result.get("rule", {})
            val = rule_res.get(code, {})
            total_checks += 1
            if not val:
                skip_count += 1
            elif val.get("pass"):
                pass_count += 1
            else:
                fail_count += 1
                failure_counter[code] += 1
                category_failures["Rule-based"][code] += 1

        # LLM Story
        for code in LLM_STORY_CODES:
            llm_res = result.get("llm_story", {})
            val = llm_res.get(code, {})
            total_checks += 1
            if not val:
                skip_count += 1
            elif val.get("pass"):
                pass_count += 1
            else:
                fail_count += 1
                failure_counter[code] += 1
                category_failures["LLM-Story"][code] += 1

        # LLM Text
        for code in LLM_TEXT_CODES:
            llm_text_res = result.get("llm_text", {})
            val = llm_text_res.get(code, {})
            total_checks += 1
            if not val:
                skip_count += 1
            elif val.get("pass"):
                pass_count += 1
            else:
                fail_count += 1
                failure_counter[code] += 1
                category_failures["LLM-Text"][code] += 1

    return {
        "overall_pass_rate": pass_count / (pass_count + fail_count) if (pass_count + fail_count) > 0 else 0,
        "total_checks": total_checks,
        "pass_count": pass_count,
        "fail_count": fail_count,
        "skip_count": skip_count,
        "failures_by_code": dict(failure_counter),
        "failures_by_category": {
            cat: dict(failures)
            for cat, failures in category_failures.items()
        }
    }


def generate_analysis_sheet(wb, all_results):
    """
    Excel에 분석 시트 추가
    """
    analytics = analyze_failures(all_results)
    ws = wb.create_sheet("분석", 1)  # "상세 결과" 다음에 추가

    # 전체 통계
    ws["A1"] = "QA 분석 대시보드"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "전체 통계"
    ws["A3"].font = Font(bold=True)

    row = 4
    ws[f"A{row}"] = "총 검증 항목"
    ws[f"B{row}"] = analytics["total_checks"]

    row += 1
    ws[f"A{row}"] = "PASS"
    ws[f"B{row}"] = analytics["pass_count"]
    ws[f"B{row}"].fill = PatternFill("solid", fgColor="C6EFCE")

    row += 1
    ws[f"A{row}"] = "FAIL"
    ws[f"B{row}"] = analytics["fail_count"]
    ws[f"B{row}"].fill = PatternFill("solid", fgColor="FFC7CE")

    row += 1
    ws[f"A{row}"] = "SKIP"
    ws[f"B{row}"] = analytics["skip_count"]

    row += 1
    ws[f"A{row}"] = "PASS율"
    ws[f"B{row}"] = f"{analytics['overall_pass_rate']*100:.1f}%"

    # 항목별 FAIL 분석
    row += 3
    ws[f"A{row}"] = "항목별 FAIL 빈도"
    ws[f"A{row}"].font = Font(bold=True)

    row += 1
    ws[f"A{row}"] = "항목"
    ws[f"B{row}"] = "FAIL 횟수"
    for c in ["A", "B"]:
        ws[f"{c}{row}"].font = Font(bold=True)
        ws[f"{c}{row}"].fill = PatternFill("solid", fgColor="D3D3D3")

    # FAIL 횟수 정렬
    sorted_failures = sorted(
        analytics["failures_by_code"].items(),
        key=lambda x: x[1],
        reverse=True
    )

    row += 1
    for code, count in sorted_failures[:15]:  # 상위 15개
        ws[f"A{row}"] = code
        ws[f"B{row}"] = count
        ws[f"B{row}"].fill = PatternFill("solid", fgColor="FFE6E6")
        row += 1

    # 카테고리별 분석
    row += 2
    ws[f"A{row}"] = "카테고리별 분석"
    ws[f"A{row}"].font = Font(bold=True)

    row += 1
    ws[f"A{row}"] = "카테고리"
    ws[f"B{row}"] = "FAIL 수"
    ws[f"C{row}"] = "비율"
    for c in ["A", "B", "C"]:
        ws[f"{c}{row}"].font = Font(bold=True)
        ws[f"{c}{row}"].fill = PatternFill("solid", fgColor="D3D3D3")

    row += 1
    for cat in ["Rule-based", "LLM-Story", "LLM-Text"]:
        failures = analytics["failures_by_category"].get(cat, {})
        fail_count = sum(failures.values())
        total_by_cat = len(RULE_CODES) if cat == "Rule-based" else (
            len(LLM_STORY_CODES) if cat == "LLM-Story" else len(LLM_TEXT_CODES)
        )
        total_by_cat *= len([r for r in all_results if r.get("turn") == 1])

        ws[f"A{row}"] = cat
        ws[f"B{row}"] = fail_count
        ws[f"C{row}"] = f"{fail_count}/{total_by_cat}"
        row += 1

    # 열 너비 조정
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15


def print_analysis(all_results):
    """콘솔에 분석 결과 출력"""
    analytics = analyze_failures(all_results)

    print("\n" + "="*60)
    print("📊 QA 분석 결과")
    print("="*60)

    print(f"\n✓ 총 검증 항목: {analytics['total_checks']}")
    print(f"  • PASS: {analytics['pass_count']} ({analytics['pass_count']/analytics['total_checks']*100:.1f}%)")
    print(f"  • FAIL: {analytics['fail_count']} ({analytics['fail_count']/analytics['total_checks']*100:.1f}%)")
    print(f"  • SKIP: {analytics['skip_count']}")
    print(f"\n📈 전체 PASS율: {analytics['overall_pass_rate']*100:.1f}%")

    print(f"\n🔴 상위 FAIL 항목:")
    sorted_failures = sorted(
        analytics["failures_by_code"].items(),
        key=lambda x: x[1],
        reverse=True
    )
    for code, count in sorted_failures[:10]:
        print(f"  • {code}: {count}회")

    print(f"\n📂 카테고리별 분석:")
    for cat in ["Rule-based", "LLM-Story", "LLM-Text"]:
        failures = analytics["failures_by_category"].get(cat, {})
        fail_count = sum(failures.values())
        print(f"  • {cat}: {fail_count}회 FAIL")

    print("\n" + "="*60 + "\n")
